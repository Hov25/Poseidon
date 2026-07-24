# -*- coding: utf-8 -*-
"""
N360 Import File Audit Tool - core logic (cloud/Streamlit port)
Parses .imp import files and exports a formatted Excel billing audit report.

This module contains only the pure parsing/export logic from the original
desktop tool (neptune360_audit_tool.py) - no tkinter, no local file dialogs.
The Streamlit app (app.py) provides the web UI on top of these functions.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
OPENPYXL_OK = True


# --- Parser ------------------------------------------------------------------

def parse_imp_file(filepath: str) -> tuple[list[dict], dict]:
    """Parse a Neptune 360 .imp file and return records + metadata."""
    records = []
    current_prm = None
    current_mtr = None
    meta = {"company": "", "route": "", "read_date": "", "total_lines": 0}

    with open(filepath, "r", encoding="latin-1") as f:
        lines = f.readlines()

    meta["total_lines"] = len(lines)

    for line in lines:
        line = line.rstrip("\r\n")
        if len(line) < 5:
            continue
        rec = line[0:5]

        if rec == "COMHD":
            meta["company"] = line[5:9].strip()

        elif rec == "RTEHD":
            meta["route"] = line[13:23].strip()
            meta["read_date"] = line[23:31].strip()

        elif rec == "PRMDT":
            current_prm = {
                "address":        line[5:31].strip(),
                "customer_name":  line[57:83].strip(),
                "account_number": line[103:123].strip(),
                "account_status": line[123:127].strip(),
                # GPS - adjust byte offsets below if your .imp version differs
                "latitude":       line[183:197].strip() if len(line) > 183 else "",
                "longitude":      line[197:211].strip() if len(line) > 197 else "",
            }

        elif rec == "PRMD2":
            current_prm = {
                "address":        (line[158:183].strip() or line[5:25].strip()),
                "customer_name":  line[25:51].strip(),
                "account_number": line[103:123].strip(),
                "account_status": line[123:127].strip(),
                # GPS - adjust byte offsets below if your .imp version differs
                "latitude":       line[183:197].strip() if len(line) > 183 else "",
                "longitude":      line[197:211].strip() if len(line) > 197 else "",
            }

        elif rec == "MTRDT":
            raw_size = line[85:93].strip()
            current_mtr = {
                "meter_number":        line[37:57].strip(),
                "meter_type":          line[77:81].strip() if len(line) > 77 else "",
                "meter_size":          raw_size or "5/8\"",
                "meter_size_missing":  not raw_size,
                "meter_uom":           line[107:110].strip(),
            }

        elif rec == "RDGDT" and current_prm and current_mtr:
            collection_id  = line[9:22].strip()
            dials          = line[49:51].strip()
            decimals       = line[53:55].strip()
            prev_read_raw  = line[78:88].strip()
            hi_limit_raw   = line[58:68].strip()
            low_limit_raw  = line[68:78].strip()

            try:
                pr  = int(prev_read_raw)
                dec = int(decimals)
                prev_read = f"{pr / (10**dec):.{dec}f}" if dec > 0 else str(pr)
            except Exception:
                prev_read = prev_read_raw

            try:
                hi = int(hi_limit_raw)
                lo = int(low_limit_raw)
            except Exception:
                hi, lo = 0, 0

            records.append({
                **current_prm,
                **current_mtr,
                "collection_id": collection_id,
                "dials":         dials,
                "decimals":      decimals,
                "prev_read":     prev_read,
                "hi_limit":      hi,
                "low_limit":     lo,
            })

    return records, meta


# --- Expected dial count by meter size ----------------------------------------
# Neptune 360 standard: 5/8"-1" = 4 dials, 1-1/2"-4" = 5 dials,
# 6"-12" = 6 dials, 16"/20" = 7 dials.

EXPECTED_DIALS = {
    '5/8"': 4, '3/4"': 4, '1"': 4,
    '1-1/2"': 5, '2"': 5, '3"': 5, '4"': 5,
    '6"': 6, '8"': 6, '10"': 6, '12"': 6,
    '16"': 7, '20"': 7,
}

# Common alternate spellings/formats seen in .imp files map to the
# canonical keys above.
_SIZE_ALIASES = {
    '5/8': '5/8"', '5/8"': '5/8"',
    '3/4': '3/4"', '3/4"': '3/4"',
    '1': '1"', '1"': '1"',
    '1-1/2': '1-1/2"', '1-1/2"': '1-1/2"',
    '1 1/2': '1-1/2"', '1 1/2"': '1-1/2"',
    '1.5': '1-1/2"', '1.5"': '1-1/2"',
    '2': '2"', '2"': '2"',
    '3': '3"', '3"': '3"',
    '4': '4"', '4"': '4"',
    '6': '6"', '6"': '6"',
    '8': '8"', '8"': '8"',
    '10': '10"', '10"': '10"',
    '12': '12"', '12"': '12"',
    '16': '16"', '16"': '16"',
    '20': '20"', '20"': '20"',
}


def _normalize_size(size: str) -> str:
    """Map a raw meter-size string to its canonical key, or '' if unrecognized."""
    if not size:
        return ""
    key = size.strip()
    return _SIZE_ALIASES.get(key, "")


def get_expected_dials(size: str):
    """Return the expected dial count for a meter size, or None if unrecognized."""
    key = _normalize_size(size)
    return EXPECTED_DIALS.get(key) if key else None


def compute_dial_comparison(records: list[dict]) -> list[dict]:
    """For each record, compare its actual dial count to the expected count
    for its meter size. Adds 'expected_dials' and 'dial_status' keys, where
    dial_status is one of: 'Match', 'Mismatch', 'Missing Dials', 'Unknown Size'.
    """
    out = []
    for r in records:
        size = r.get("meter_size", "")
        dials_raw = (r.get("dials") or "").strip()
        expected = get_expected_dials(size)

        actual = None
        if dials_raw:
            try:
                actual = int(dials_raw)
            except ValueError:
                actual = None

        if expected is None:
            status = "Unknown Size"
        elif actual is None:
            status = "Missing Dials"
        elif actual == expected:
            status = "Match"
        else:
            status = "Mismatch"

        out.append({
            **r,
            "expected_dials": expected,
            "actual_dials": actual,
            "dial_status": status,
        })
    return out


def export_dial_comparison_xlsx(records: list[dict], meta: dict, out_path: str) -> None:
    """Export a standalone Excel report of the dial-count comparison,
    with mismatches/unknowns highlighted."""
    comparison = compute_dial_comparison(records)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dial Value Check"

    hdr_font  = Font(name="Calibri", bold=True, color=WHITE, size=10)
    hdr_fill  = PatternFill("solid", start_color=NAVY)
    data_font = Font(name="Calibri", size=10)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")
    bdr       = _border()

    headers = [
        ("Account Number", 16), ("Customer Name", 26), ("Meter Number", 18),
        ("Collection ID", 18), ("Meter Size", 12), ("Expected Dials", 15),
        ("Actual Dials", 13), ("Status", 16),
    ]
    total_cols = len(headers)
    last_col = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "N360 Import File Audit Tool  -  Dial Value Check"
    ws["A1"].font = Font(name="Calibri", bold=True, size=15, color=NAVY)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    mismatches = sum(1 for c in comparison if c["dial_status"] == "Mismatch")
    unknowns   = sum(1 for c in comparison if c["dial_status"] == "Unknown Size")
    missing    = sum(1 for c in comparison if c["dial_status"] == "Missing Dials")

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        f"Total Records: {len(comparison)}   |   Mismatches: {mismatches}   |   "
        f"Unknown Sizes: {unknowns}   |   Missing Dials: {missing}"
    )
    ws["A2"].font = Font(name="Calibri", bold=True, size=10, color="7B2D00")
    ws["A2"].fill = PatternFill("solid", start_color="FCE4D6")
    ws["A2"].alignment = left
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    for ci, (label, width) in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=label)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[4].height = 20

    fill_match    = PatternFill("solid", start_color=GREEN)
    fill_mismatch = PatternFill("solid", start_color=RED)
    fill_unknown  = PatternFill("solid", start_color="D9D9D9")
    fill_missing  = PatternFill("solid", start_color=YELLOW)

    status_fills = {
        "Match": fill_match, "Mismatch": fill_mismatch,
        "Unknown Size": fill_unknown, "Missing Dials": fill_missing,
    }

    for ri, rec in enumerate(comparison, 5):
        vals = [
            rec["account_number"], rec["customer_name"], rec["meter_number"],
            rec["collection_id"], rec["meter_size"],
            rec["expected_dials"] if rec["expected_dials"] is not None else "--",
            rec["actual_dials"] if rec["actual_dials"] is not None else "--",
            rec["dial_status"],
        ]
        row_fill = status_fills.get(rec["dial_status"], fill_match)
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = data_font; c.border = bdr
            c.fill = row_fill
            c.alignment = center if ci != 2 else left
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_col}{len(comparison) + 4}"

    wb.save(out_path)


# --- Data-quality helpers -----------------------------------------------------

def _compute_quality(records: list[dict]) -> dict:
    """Return counts of common data-quality issues across all records."""
    from collections import Counter, defaultdict

    coll_ids   = [r["collection_id"] for r in records]
    meter_nums = [r["meter_number"]  for r in records]

    # meter sizes that appear with more than one distinct dial value
    size_to_dials: dict = defaultdict(set)
    for r in records:
        if r["meter_size"] and r["dials"]:
            size_to_dials[r["meter_size"]].add(r["dials"])
    conflict_sizes = {sz for sz, dv in size_to_dials.items() if len(dv) > 1}

    cc = Counter(coll_ids)
    mc = Counter(meter_nums)

    return {
        "conflict_sizes":       conflict_sizes,
        "meter_dial_conflicts": sum(1 for r in records if r["meter_size"] in conflict_sizes),
        "missing_dials":        sum(1 for r in records if not r["dials"]),
        "missing_meter_sz":     sum(1 for r in records if r.get("meter_size_missing")),
        "dup_coll_ids":         sum(1 for v in coll_ids   if v and cc[v] > 1),
        "dup_meter_nums":       sum(1 for v in meter_nums if mc[v] > 1),
    }


# --- Data-quality sheet -------------------------------------------------------

def _export_quality_sheet(wb, records: list[dict], dq: dict) -> None:
    """Write one tab per data-quality issue type."""
    from collections import Counter, defaultdict

    # -- Shared style factory -----------------------------------------------
    navy_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    navy_fill   = PatternFill("solid", start_color="1F4E79")
    title_font  = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    data_font   = Font(name="Calibri", size=10)
    orange_fill = PatternFill("solid", start_color="FCE4D6")
    alt_fill    = PatternFill("solid", start_color="EBF3FB")
    white_fill  = PatternFill("solid", start_color="FFFFFF")
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")
    bdr         = _border()

    def make_sheet(title):
        ws = wb.create_sheet(title)
        ws["A1"] = f"N360 Import File Audit Tool  -  {title}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 6   # spacer
        return ws

    def col_headers(ws, row, labels, widths):
        for ci, (lbl, w) in enumerate(zip(labels, widths), 1):
            c = ws.cell(row=row, column=ci, value=lbl)
            c.font = navy_font; c.fill = navy_fill
            c.alignment = center; c.border = bdr
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = 18
        ws.freeze_panes = f"A{row + 1}"

    def data_row(ws, row, vals, highlight=False):
        fill = orange_fill if highlight else (alt_fill if row % 2 == 0 else white_fill)
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = data_font; c.fill = fill; c.border = bdr
            c.alignment = left if ci == 2 else center
        ws.row_dimensions[row].height = 16

    # -- Tab 1: Meter Size / Dial Conflicts --------------------------------
    conflict_sizes = dq["conflict_sizes"]
    ws1 = make_sheet("Meter Size-Dial Conflicts")
    col_headers(ws1, 3,
        ["Meter Size", "All Dial Values for Size", "This Account's Dial", "Account Number",
         "Customer Name", "Meter Number", "Collection ID"],
        [12, 24, 20, 18, 26, 18, 18])

    size_groups: dict = defaultdict(list)
    for r in records:
        if r["meter_size"] in conflict_sizes:
            size_groups[r["meter_size"]].append(r)

    ri = 4
    for sz in sorted(size_groups):
        grp = size_groups[sz]
        all_dials = ", ".join(sorted({r["dials"] for r in grp if r["dials"]}))
        for r in grp:
            data_row(ws1, ri, [sz, all_dials, r["dials"], r["account_number"],
                               r["customer_name"], r["meter_number"], r["collection_id"]],
                     highlight=True)
            ri += 1

    # -- Tab 2: Meter Size-Dial Conflict Count ----------------------------
    ws_cnt = make_sheet("Dials Conflict Count")
    col_headers(ws_cnt, 3,
        ["Meter Size", "Dial Value", "Account Count", "% of Size Group"],
        [14, 14, 16, 18])

    green_fill  = PatternFill("solid", start_color="E2EFDA")
    yellow_fill = PatternFill("solid", start_color="FFF2CC")
    size_fills  = [alt_fill, green_fill, yellow_fill,
                   PatternFill("solid", start_color="D9E1F2")]

    ri = 4
    for si, sz in enumerate(sorted(size_groups)):
        grp       = size_groups[sz]
        sz_total  = len(grp)
        dial_counts: dict = {}
        for r in grp:
            d = r["dials"] if r["dials"] else "(blank)"
            dial_counts[d] = dial_counts.get(d, 0) + 1
        row_fill = size_fills[si % len(size_fills)]
        for dial in sorted(dial_counts):
            count = dial_counts[dial]
            pct   = f"{count / sz_total * 100:.1f}%" if sz_total else "--"
            for ci, val in enumerate([sz, dial, count, pct], 1):
                c = ws_cnt.cell(row=ri, column=ci, value=val)
                c.font = data_font; c.fill = row_fill; c.border = bdr
                c.alignment = center
            ws_cnt.row_dimensions[ri].height = 16
            ri += 1
        # Subtotal row per meter size
        total_font = Font(name="Calibri", size=10, bold=True)
        for ci, val in enumerate([sz, "TOTAL", sz_total, "100%"], 1):
            c = ws_cnt.cell(row=ri, column=ci, value=val)
            c.font = total_font
            c.fill = PatternFill("solid", start_color="1F4E79")
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.border = bdr; c.alignment = center
        ws_cnt.row_dimensions[ri].height = 17
        ri += 1

    # -- Tab 3: Duplicate Collection IDs -----------------------------------
    cc = Counter(r["collection_id"] for r in records)
    dup_coll = {cid for cid, cnt in cc.items() if cnt > 1 and cid}
    ws2 = make_sheet("Duplicate Collection IDs")
    col_headers(ws2, 3,
        ["Collection ID", "Account Number", "Customer Name", "Meter Number", "Address"],
        [18, 18, 26, 18, 28])
    ri = 4
    for r in [r for r in records if r["collection_id"] in dup_coll]:
        data_row(ws2, ri, [r["collection_id"], r["account_number"], r["customer_name"],
                           r["meter_number"], r["address"]])
        ri += 1

    # -- Tab 4: Duplicate Meter Numbers ------------------------------------
    mc = Counter(r["meter_number"] for r in records)
    dup_mtrs = {mn for mn, cnt in mc.items() if cnt > 1}
    ws3 = make_sheet("Duplicate Meter Numbers")
    col_headers(ws3, 3,
        ["Meter Number", "Account Number", "Customer Name", "Collection ID", "Address"],
        [18, 18, 26, 18, 28])
    ri = 4
    for r in [r for r in records if r["meter_number"] in dup_mtrs]:
        data_row(ws3, ri, [r["meter_number"], r["account_number"], r["customer_name"],
                           r["collection_id"], r["address"]])
        ri += 1

    # -- Tab 5: Missing Meter Sizes -----------------------------------------
    ws5 = make_sheet("Missing Meter Sizes")
    col_headers(ws5, 3,
        ["Account Number", "Customer Name", "Meter Number", "Collection ID", "Dials"],
        [18, 26, 18, 18, 8])
    ri = 4
    for r in [r for r in records if r.get("meter_size_missing")]:
        data_row(ws5, ri, [r["account_number"], r["customer_name"], r["meter_number"],
                           r["collection_id"], r["dials"]])
        ri += 1

    # -- Tab 6: Meter Size / Dial Reference --------------------------------
    # Full lookup table - every recognised Neptune 360 meter size with its
    # expected dial count and the notes column used in this file, so the
    # auditor never has to leave the workbook to verify a reading.

    DIAL_REFERENCE = [
        # (Meter Size,  Expected Dials,  Meter Class,           Notes)
        ('5/8"',   4, 'Small Residential',          'Standard residential service'),
        ('3/4"',   4, 'Small Residential',          'Standard residential service'),
        ('1"',     4, 'Residential / Small Comm.',  'Small commercial or multi-family'),
        ('1-1/2"', 5, 'Commercial',                 'Light commercial'),
        ('2"',     5, 'Commercial',                 'Commercial service'),
        ('3"',     5, 'Commercial',                 'Medium commercial / light industrial'),
        ('4"',     5, 'Commercial',                 'Medium commercial / light industrial'),
        ('6"',     6, 'Large Commercial / Indust.', 'Large commercial or industrial'),
        ('8"',     6, 'Large Commercial / Indust.', 'Large commercial or industrial'),
        ('10"',    6, 'Large Commercial / Indust.', 'Large commercial or industrial'),
        ('12"',    6, 'Large Commercial / Indust.', 'Large commercial or industrial'),
        ('16"',    7, 'Transmission / Bulk',        'Transmission main or bulk metering'),
        ('18"',    7, 'Transmission / Bulk',        'Transmission main or bulk metering'),
    ]

    # Group by expected-dials band for colour banding
    BAND_FILLS = {
        4: PatternFill("solid", start_color="E2EFDA"),   # green  - small
        5: PatternFill("solid", start_color="EBF3FB"),   # blue   - commercial
        6: PatternFill("solid", start_color="FFF2CC"),   # yellow - large
        7: PatternFill("solid", start_color="FCE4D6"),   # orange - bulk
    }
    BAND_LABELS = {
        4: "4 Dials  (5/8\" - 1\")",
        5: "5 Dials  (1-1/2\" - 4\")",
        6: "6 Dials  (6\" - 12\")",
        7: "7 Dials  (16\" - 18\")",
    }

    ws6 = make_sheet("Meter Size-Dial Reference")

    # Subtitle description
    ws6["A2"] = (
        "Neptune 360 standard expected dial counts by meter size.  "
        "Use this table to verify that dials in the import file match the installed meter."
    )
    ws6["A2"].font      = Font(name="Calibri", size=9, italic=True, color="555555")
    ws6["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws6.row_dimensions[2].height = 22
    ws6.row_dimensions[3].height = 6   # spacer row between subtitle and headers

    col_headers(ws6, 4,
        ["Meter Size", "Expected Dials", "Dial Band", "Meter Class", "Notes"],
        [13, 16, 26, 26, 36])

    # Band-summary rows inserted before each group
    last_band = None
    ri = 5
    for (size, dials, cls, notes) in DIAL_REFERENCE:
        band_fill = BAND_FILLS[dials]

        # Insert a band-header row whenever the dial group changes
        if dials != last_band:
            last_band = dials
            band_label = BAND_LABELS[dials]
            for ci in range(1, 6):
                c = ws6.cell(row=ri, column=ci,
                             value=band_label if ci == 1 else "")
                c.font      = Font(name="Calibri", bold=True, size=10, color="1F4E79")
                c.fill      = PatternFill("solid", start_color="D9E1F2")
                c.border    = bdr
                c.alignment = Alignment(horizontal="left" if ci == 1 else "center",
                                        vertical="center")
            ws6.row_dimensions[ri].height = 17
            ri += 1

        for ci, val in enumerate([size, dials, BAND_LABELS[dials], cls, notes], 1):
            c = ws6.cell(row=ri, column=ci, value=val)
            c.font      = data_font
            c.fill      = band_fill
            c.border    = bdr
            c.alignment = Alignment(
                horizontal="center" if ci in (1, 2) else "left",
                vertical="center"
            )
        ws6.row_dimensions[ri].height = 16
        ri += 1

    # Colour-key legend below the table
    ri += 1
    ws6.cell(row=ri, column=1, value="Colour Key").font = \
        Font(name="Calibri", bold=True, size=10, color="1F4E79")
    ri += 1
    for dials, label in BAND_LABELS.items():
        c = ws6.cell(row=ri, column=1, value=label)
        c.font   = Font(name="Calibri", size=10)
        c.fill   = BAND_FILLS[dials]
        c.border = bdr
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws6.row_dimensions[ri].height = 15
        ri += 1


# --- Excel exporter -----------------------------------------------------------

HEADERS = [
    ("Account Number",  16),
    ("Customer Name",   26),
    ("Address",         24),
    ("Status",          10),
    ("Meter Number",    18),
    ("Collection ID",   18),
    ("Meter Size",      11),
    ("Meter Type",      18),
    ("UOM",              8),
    ("Dials",            7),
    ("Decimals",         9),
    ("Previous Read",   14),
    ("Hi Limit",        13),
    ("Low Limit",       13),
    ("Latitude",        14),
    ("Longitude",       14),
]

NAVY   = "1F4E79"
WHITE  = "FFFFFF"
BLUE1  = "D6E4F0"
BLUE2  = "EBF3FB"
GREEN  = "E2EFDA"
YELLOW = "FFF2CC"
RED    = "FCE4D6"


def _side(color="AAAAAA"):
    return Side(style="thin", color=color)


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def export_xlsx(records: list[dict], meta: dict, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Billing Audit"

    hdr_font  = Font(name="Calibri", bold=True, color=WHITE, size=10)
    hdr_fill  = PatternFill("solid", start_color=NAVY)
    data_font = Font(name="Calibri", size=10)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left      = Alignment(horizontal="left",   vertical="center")
    bdr       = _border()

    # -- Title block --------------------------------------------------------
    total_cols = len(HEADERS)
    last_col   = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "N360 Import File Audit Tool  -  Billing Audit Report"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=15, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    rd = meta.get("read_date", "")
    if len(rd) == 8:
        rd = f"{rd[:4]}-{rd[4:6]}-{rd[6:]}"

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        f"Company: {meta.get('company','')}   |   "
        f"Route: {meta.get('route','')}   |   "
        f"Read Date: {rd}   |   "
        f"Total Records: {len(records)}"
    )
    ws["A2"].font      = Font(name="Calibri", italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # -- Summary row 3 - account totals -------------------------------------
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = (
        f"Total Accounts: {len(records)}   |   "
        f"Active: {sum(1 for r in records if r['account_status'] in ('ACTI','AWZ'))}   |   "
        f"Inactive: {sum(1 for r in records if r['account_status'] in ('INAC','IWU'))}"
    )
    ws["A3"].font      = Font(name="Calibri", bold=True, size=10, color=NAVY)
    ws["A3"].fill      = PatternFill("solid", start_color="D9E1F2")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18

    # -- Summary row 4 - data-quality flags --------------------------------
    dq = _compute_quality(records)
    ws.merge_cells(f"A4:{last_col}4")
    ws["A4"] = (
        f"Data Quality -   "
        f"Meter Size/Dial Conflicts: {dq['meter_dial_conflicts']}   |   "
        f"Missing Meter Sizes: {dq['missing_meter_sz']}   |   "
        f"Duplicate Collection IDs: {dq['dup_coll_ids']}   |   "
        f"Duplicate Meter Numbers: {dq['dup_meter_nums']}"
    )
    ws["A4"].font      = Font(name="Calibri", bold=True, size=10, color="7B2D00")
    ws["A4"].fill      = PatternFill("solid", start_color="FCE4D6")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 18

    ws.row_dimensions[5].height = 6   # spacer

    # -- Column headers -----------------------------------------------------
    for ci, (label, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=6, column=ci, value=label)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = bdr
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[6].height = 20

    # -- Pre-build fill palette (avoids re-creating objects per cell) -------
    fill_even   = PatternFill("solid", start_color=BLUE1)
    fill_odd    = PatternFill("solid", start_color=BLUE2)
    fill_green  = PatternFill("solid", start_color=GREEN)
    fill_yellow = PatternFill("solid", start_color=YELLOW)
    fill_red    = PatternFill("solid", start_color=RED)

    # -- Data rows ----------------------------------------------------------
    for ri, rec in enumerate(records, 7):
        base_fill = fill_even if ri % 2 == 0 else fill_odd

        status = rec["account_status"]
        status_fill = (
            fill_green  if status in ("ACTI", "AWZ") else
            fill_yellow if status == "IWU"           else
            fill_red
        )

        row_vals = [
            rec["account_number"],
            rec["customer_name"],
            rec["address"],
            status,
            rec["meter_number"],
            rec["collection_id"],
            rec["meter_size"],
            rec.get("meter_type", ""),
            rec["meter_uom"],
            rec["dials"],
            rec["decimals"],
            rec["prev_read"],
            rec["hi_limit"],
            rec["low_limit"],
            rec.get("latitude",  ""),
            rec.get("longitude", ""),
        ]

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = data_font
            cell.border    = bdr
            # Status column gets its own colour; numeric cols centred
            if ci == 4:
                cell.fill      = status_fill
                cell.alignment = center
            elif ci in (1, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16):
                cell.fill      = base_fill
                cell.alignment = center
            else:
                cell.fill      = base_fill
                cell.alignment = left

        ws.row_dimensions[ri].height = 16

    # -- Freeze & filter ----------------------------------------------------
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:{last_col}{len(records) + 6}"

    # -- Data Quality sheet -------------------------------------------------
    _export_quality_sheet(wb, records, dq)

    wb.save(out_path)

