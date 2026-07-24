"""
MIU System Assessment Tool — core logic (cloud/Streamlit port)
Parses Neptune 360 .imp files, .xlsx, and .csv files to identify MIU type
and estimate MIU age.

This module contains only the pure parsing/lookup/export logic from the
original desktop tool (miu_assessment_tool.py) — no tkinter, no local file
dialogs. The Streamlit app (app.py) provides the web UI on top of these
functions.
"""

import csv as _csv
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
OPENPYXL_OK = True


CURRENT_YEAR = datetime.now().year


# ─── MIU Range Lookup Table ───────────────────────────────────────────────────
# Source: "Updated MIU by Year and model.csv"
# NOTE: The 2100M Gas Remote range must appear before the broader 2000M Gas
#       range so the more-specific entry wins on overlap.

MIU_RANGES = [
    # ── Legacy Cellular ───────────────────────────────────────────────────────
    (400000000,  499999999,  "Legacy Cellular", "CMIU Verizon",                2017, 2017),
    (500000000,  599999999,  "Legacy Cellular", "CMIU AT&T",                   2017, 2017),

    # ── R450 (sorted by start ID) ─────────────────────────────────────────────
    (110000000,  111082515,  "R450", "R450", 2012, 2012),
    (111082516,  112131332,  "R450", "R450", 2013, 2013),
    (112131333,  112527102,  "R450", "R450", 2014, 2014),
    (112527103,  112815048,  "R450", "R450", 2015, 2015),
    (112815049,  114044970,  "R450", "R450", 2016, 2016),
    (114044971,  114197882,  "R450", "R450", 2017, 2017),
    (114197883,  114387746,  "R450", "R450", 2018, 2018),
    (114387747,  114500158,  "R450", "R450", 2019, 2019),
    (114500159,  114534096,  "R450", "R450", 2020, 2020),
    (114534097,  114593748,  "R450", "R450", 2021, 2021),
    (114593749,  114590805,  "R450", "R450", 2022, 2022),
    (114590806,  114633311,  "R450", "R450", 2023, 2023),
    (114633312,  119999999,  "R450", "R450", 2024, 2024),

    # ── R900 (original / 1998) ────────────────────────────────────────────────
    (1000000000, 1199999999, "R900", "R900",                1998, 1998),

    # ── R900v2 ────────────────────────────────────────────────────────────────
    (1400010000, 1439999999, "R900", "R900v2",               2003, 2003),

    # ── R900v3 ────────────────────────────────────────────────────────────────
    (1440000000, 1459999999, "R900", "R900v3",               2004, 2004),
    (1460000000, 1469999999, "R900", "R900v3",               2004, 2004),
    (1470000000, 1479999999, "R900", "R900v3",               2004, 2004),
    (1490000000, 1499999999, "R900", "R900v3",               2007, 2007),
    (1480000000, 1489999999, "R900", "R900v3",               2008, 2008),

    # ── R900iv3 ───────────────────────────────────────────────────────────────
    (1810000000, 1819999999, "R900", "R900iv3",              2005, 2005),
    (1820000000, 1829999999, "R900", "R900iv3",              2007, 2007),

    # ── R900iv3 Datalogger ────────────────────────────────────────────────────
    (1830000000, 1839999999, "R900", "R900iv3 Datalogger",   2008, 2008),
    (1840000000, 1849999999, "R900", "R900iv3 Datalogger",   2008, 2008),
    (1850000000, 1859999999, "R900", "R900iv3 Datalogger",   2013, 2013),

    # ── R900 Gas (2100M checked first — overlaps with the 2000M range below) ──
    (2100000000, 2199999999, "R900", "R900 Gas Remote (High Power)", 2007, 2007),
    (2000000000, 2099999999, "R900", "R900 Gas (High Power)",        2008, 2008),

    # ── R900 V4 (year-specific ID bands) ─────────────────────────────────────
    (1540000000, 1544092999, "R900", "R900 V4", 2015, 2015),
    (1544093000, 1548186199, "R900", "R900 V4", 2016, 2016),
    (1548186200, 1552279199, "R900", "R900 V4", 2017, 2017),
    (1552279200, 1556372199, "R900", "R900 V4", 2018, 2018),
    (1556372200, 1560465199, "R900", "R900 V4", 2019, 2019),
    (1560465200, 1564558399, "R900", "R900 V4", 2020, 2020),
    (1564558400, 1568651399, "R900", "R900 V4", 2021, 2021),
    (1568651400, 1572744399, "R900", "R900 V4", 2022, 2022),
    (1572744400, 1576837399, "R900", "R900 V4", 2023, 2023),
    (1576837400, 1582509197, "R900", "R900 V4", 2024, 2024),
    (1582509198, 1584833319, "R900", "R900 V4", 2025, 2025),
    (1584833320, 1599999999, "R900", "R900 V4", 2026, 2026),

    # ── R900v5 / R900iv5 ──────────────────────────────────────────────────────
    (700000000,  799999999,  "R900", "R900v5 / R900iv5",       2018, 2018),

    # ── R900 Cellular Endpoint ────────────────────────────────────────────────
    (220000000,  239999999,  "R900", "R900 Cellular Endpoint",          2021, 2021),

    # ── R900 Multi Carrier Cellular Endpoint ──────────────────────────────────
    (320000000,  329999999,  "R900", "R900 Multi Carrier Cellular Endpoint", 2025, 2025),
]


def lookup_miu(collection_id: str) -> tuple:
    """
    Classify an MIU by matching its collection ID against known ID ranges.
    Returns (system, miu_type, start_year, end_year, serial_status).
    serial_status is one of: "ok", "missing", "invalid".
    """
    stripped = (collection_id or "").strip()
    if not stripped:
        return "Missing", "Missing MIU number", None, None, "missing"
    try:
        cid = int(stripped)
    except ValueError:
        return "Invalid", "Invalid Serial Number", None, None, "invalid"

    for lo, hi, system, miu_type, start_yr, end_yr in MIU_RANGES:
        if lo <= cid <= hi:
            return system, miu_type, start_yr, end_yr, "ok"

    return "Invalid", "Invalid Serial Number", None, None, "invalid"


# ─── Age Helpers ──────────────────────────────────────────────────────────────

def age_label(start_year, end_year) -> str:
    if start_year is None:
        return "Unknown"
    max_age = CURRENT_YEAR - start_year
    min_age = CURRENT_YEAR - (end_year if end_year else CURRENT_YEAR)
    if min_age == max_age:
        return f"{max_age} yr{'s' if max_age != 1 else ''}"
    return f"{min_age}–{max_age} yrs"


def age_category(start_year) -> str:
    if start_year is None:
        return "Unknown"
    age = CURRENT_YEAR - start_year
    if age <= 5:
        return "New (<=5 yrs)"
    if age <= 10:
        return "Moderate (6-10 yrs)"
    if age <= 15:
        return "Aging (11-15 yrs)"
    return "End of Life (>15 yrs)"


def age_fill_color(category: str) -> str:
    return {
        "New (<=5 yrs)":         "E2EFDA",
        "Moderate (6-10 yrs)":   "FFF2CC",
        "Aging (11-15 yrs)":     "FCE4D6",
        "End of Life (>15 yrs)": "F4CCCC",
    }.get(category, "FFFFFF")


# ─── .imp Parser ──────────────────────────────────────────────────────────────

def parse_imp_file(filepath: str) -> tuple:
    """Parse a Neptune 360 .imp file; return MIU records + metadata."""
    records = []
    current_prm = None
    current_mtr = None
    meta = {"company": "", "route": "", "read_date": "", "total_lines": 0,
            "file_type": "imp"}

    with open(filepath, "r", encoding="latin-1") as f:
        lines = f.readlines()

    meta["total_lines"] = len(lines)

    for line in lines:
        line = line.rstrip("\r\n")
        if len(line) < 5:
            continue
        rec = line[:5]

        if rec == "COMHD":
            meta["company"] = line[5:9].strip()

        elif rec == "RTEHD":
            meta["route"]     = line[13:23].strip()
            meta["read_date"] = line[23:31].strip()

        elif rec == "PRMDT":
            current_prm = {
                "address":        line[5:31].strip(),
                "customer_name":  line[57:83].strip(),
                "account_number": line[103:123].strip(),
            }

        elif rec == "PRMD2":
            current_prm = {
                "address":        (line[158:183].strip() or line[5:25].strip()),
                "customer_name":  line[25:51].strip(),
                "account_number": line[103:123].strip(),
            }

        elif rec == "MTRDT":
            current_mtr = {
                "meter_number": line[37:57].strip(),
                "meter_size":   line[85:93].strip() or "5/8\"",
            }

        elif rec == "RDGDT" and current_prm and current_mtr:
            collection_id                                       = line[9:22].strip()
            system, miu_type, start_yr, end_yr, serial_status  = lookup_miu(collection_id)

            records.append({
                "account_number": current_prm["account_number"],
                "customer_name":  current_prm["customer_name"],
                "address":        current_prm["address"],
                "meter_number":   current_mtr["meter_number"],
                "meter_size":     current_mtr["meter_size"],
                "miu_serial":     collection_id,
                "serial_status":  serial_status,
                "system":         system,
                "miu_type":       miu_type,
                "est_year":       start_yr,
                "end_year":       end_yr,
                "age_label":      age_label(start_yr, end_yr),
                "age_category":   age_category(start_yr),
            })

    records.sort(key=lambda r: r["est_year"] if r["est_year"] else 9999)
    return records, meta


# ─── Tabular (.xlsx / .csv) Parser ────────────────────────────────────────────

# Column name patterns that likely contain MIU serial numbers (most-specific first)
_MIU_COL_KEYWORDS = [
    ("collection", "id"), ("collection", "no"), ("collection", "#"),
    ("miu", "id"),        ("miu", "serial"),    ("miu", "number"),
    ("miu", "no"),        ("miu", "#"),
    ("serial", "number"), ("serial", "no"),     ("serial", "#"),
    ("endpoint", "id"),   ("endpoint", "no"),
    ("serial",),          ("miu",),             ("endpoint",),
]


def _detect_miu_column_index(headers: list) -> int | None:
    """Return index of the column most likely to hold MIU serial numbers."""
    hl = [str(h).lower().strip() for h in headers]
    for pattern in _MIU_COL_KEYWORDS:
        for i, h in enumerate(hl):
            if all(k in h for k in pattern):
                return i
    return None


def _normalize_serial(val) -> str:
    """Strip float suffix and whitespace that Excel sometimes adds."""
    v = str(val).strip()
    if v.endswith(".0"):
        v = v[:-2]
    return v


def _read_tabular(filepath: str) -> tuple:
    """Read .csv or .xlsx into (headers: list[str], rows: list[dict])."""
    ext = Path(filepath).suffix.lower()

    if ext == ".csv":
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = _csv.DictReader(f)
            headers = list(reader.fieldnames or [])
            rows = [dict(r) for r in reader]
        return headers, rows

    if ext in (".xlsx", ".xlsm"):
        if not OPENPYXL_OK:
            raise RuntimeError("openpyxl is required to read .xlsx files.\n"
                               "Install with:  pip install openpyxl")
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return [], []
        headers = [
            str(c).strip() if c is not None else f"Col{i + 1}"
            for i, c in enumerate(all_rows[0])
        ]
        rows = []
        for row in all_rows[1:]:
            if all(c is None for c in row):
                continue  # skip blank rows
            rows.append({
                h: (_normalize_serial(v) if v is not None else "")
                for h, v in zip(headers, row)
            })
        return headers, rows

    raise ValueError(f"Unsupported file extension: {ext}")


def parse_tabular_file(filepath: str, miu_col: str | None = None) -> tuple:
    """
    Parse an .xlsx or .csv file to identify MIU type and age.

    miu_col: override which column holds MIU serial numbers.
             If None, auto-detected.

    Returns (records, meta) in the same format as parse_imp_file().
    meta includes 'miu_column' (detected column name) and 'source_headers'.
    """
    headers, rows = _read_tabular(filepath)
    if not rows:
        return [], {
            "company": "", "route": "", "read_date": "",
            "total_lines": 0, "file_type": "tabular",
            "source_file": Path(filepath).name,
            "miu_column": None, "source_headers": headers,
        }

    # ── Resolve MIU serial column ────────────────────────────────────────────
    if miu_col and miu_col in headers:
        detected_col = miu_col
        detection_method = "user"
    else:
        idx = _detect_miu_column_index(headers)
        if idx is not None:
            detected_col = headers[idx]
            detection_method = "header"
        else:
            # Value-based fallback: find column with most serial-range hits
            best_col, best_hits = None, 0
            for col in headers:
                hits = sum(
                    1 for row in rows[:50]
                    if lookup_miu(_normalize_serial(row.get(col, "")))[4] == "ok"
                )
                if hits > best_hits:
                    best_hits, best_col = hits, col
            detected_col = best_col
            detection_method = "value" if detected_col else "none"

    # ── Helper: case-insensitive column lookup ───────────────────────────────
    def _get(row, *candidates):
        lrow = {k.lower(): v for k, v in row.items()}
        for c in candidates:
            v = lrow.get(c.lower(), "").strip()
            if v:
                return v
        return ""

    # ── Build records ────────────────────────────────────────────────────────
    records = []
    for row in rows:
        serial_raw = _normalize_serial(row.get(detected_col, "")) if detected_col else ""
        system, miu_type, start_yr, end_yr, serial_status = lookup_miu(serial_raw)

        records.append({
            "account_number": _get(row,
                "Account Number", "Account No", "Account #", "Acct No", "Acct",
                "Account", "Customer Account"),
            "customer_name":  _get(row,
                "Customer Name", "Name", "Customer", "Cust Name"),
            "address":        _get(row,
                "Address", "Service Address", "Location", "Svc Address",
                "Street Address"),
            "meter_number":   _get(row,
                "Meter Number", "Meter No", "Meter #", "Meter Serial",
                "Meter ID", "Meter"),
            "meter_size":     _get(row, "Meter Size", "Size"),
            "miu_serial":     serial_raw,
            "serial_status":  serial_status,
            "system":         system,
            "miu_type":       miu_type,
            "est_year":       start_yr,
            "end_year":       end_yr,
            "age_label":      age_label(start_yr, end_yr),
            "age_category":   age_category(start_yr),
            # Preserve original row for passthrough in tabular export
            "_raw_row":       row,
            "_headers":       headers,
        })

    records.sort(key=lambda r: r["est_year"] if r["est_year"] else 9999)

    meta = {
        "company": "", "route": "", "read_date": "",
        "total_lines": len(rows),
        "file_type": "tabular",
        "source_file": Path(filepath).name,
        "miu_column": detected_col,
        "detection_method": detection_method,
        "source_headers": headers,
    }
    return records, meta


# ─── Excel Export ─────────────────────────────────────────────────────────────

NAVY  = "1F4E79"
WHITE = "FFFFFF"


def _border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def export_xlsx(records: list, meta: dict, out_path: str) -> None:
    """Export records to a formatted Excel workbook (standard .imp layout)."""
    from collections import Counter
    wb  = Workbook()
    bdr = _border()

    hdr_font  = Font(name="Calibri", bold=True, color=WHITE, size=10)
    hdr_fill  = PatternFill("solid", start_color=NAVY)
    data_font = Font(name="Calibri", size=10)
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")

    # ── Sheet 1: MIU Age Report ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "MIU Age Report"

    rd = meta.get("read_date", "")
    if len(rd) == 8:
        rd = f"{rd[:4]}-{rd[4:6]}-{rd[6:]}"

    cats = Counter(r["age_category"] for r in records)

    ws.merge_cells("A1:L1")
    ws["A1"] = "MIU System Assessment Tool  --  MIU Age Report"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=15, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    ws["A2"] = (
        f"Company: {meta.get('company', '')}   |   "
        f"Route: {meta.get('route', '')}   |   "
        f"Read Date: {rd}   |   "
        f"Total MIUs: {len(records)}"
    )
    ws["A2"].font      = Font(name="Calibri", italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:L3")
    ws["A3"] = (
        f"New (<=5 yrs): {cats.get('New (<=5 yrs)', 0)}   |   "
        f"Moderate (6-10 yrs): {cats.get('Moderate (6-10 yrs)', 0)}   |   "
        f"Aging (11-15 yrs): {cats.get('Aging (11-15 yrs)', 0)}   |   "
        f"End of Life (>15 yrs): {cats.get('End of Life (>15 yrs)', 0)}   |   "
        f"Unknown: {cats.get('Unknown', 0)}"
    )
    ws["A3"].font      = Font(name="Calibri", bold=True, size=10, color="7B2D00")
    ws["A3"].fill      = PatternFill("solid", start_color="FCE4D6")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 6

    headers = [
        ("Account Number",  16), ("Customer Name",  26), ("Address",      24),
        ("Meter Number",    18), ("Meter Size",      11), ("Collection ID", 18),
        ("System",          16), ("MIU Type",        26), ("Mfr. Start Yr", 13),
        ("Mfr. End Yr",     13), ("Est. Age",        16), ("Age Category",  22),
    ]
    for ci, (label, width) in enumerate(headers, 1):
        c = ws.cell(row=5, column=ci, value=label)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[5].height = 20

    serial_missing_fill = PatternFill("solid", start_color="FFCCCC")
    serial_invalid_fill = PatternFill("solid", start_color="FFD7A8")
    serial_missing_font = Font(name="Calibri", size=10, bold=True, color="922B21")
    serial_invalid_font = Font(name="Calibri", size=10, bold=True, color="7E3200")

    for ri, rec in enumerate(records, 6):
        cat             = rec["age_category"]
        fill            = PatternFill("solid", start_color=age_fill_color(cat))
        end_yr_display  = rec["end_year"] if rec["end_year"] else "Present"
        serial_status   = rec.get("serial_status", "ok")
        serial_display  = rec["miu_serial"] if rec["miu_serial"] else "Missing MIU number"
        row_vals = [
            rec["account_number"], rec["customer_name"], rec["address"],
            rec["meter_number"],   rec["meter_size"],    serial_display,
            rec["system"],         rec["miu_type"],      rec["est_year"],
            end_yr_display,        rec["age_label"],     cat,
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = bdr
            c.alignment = left if ci in (2, 3) else center
            if ci == 6 and serial_status == "missing":
                c.fill = serial_missing_fill
                c.font = serial_missing_font
            elif ci == 6 and serial_status == "invalid":
                c.fill = serial_invalid_fill
                c.font = serial_invalid_font
            else:
                c.fill = fill
                c.font = data_font
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:L{len(records) + 5}"

    # ── Sheet 2: Age Summary ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Age Summary")
    _write_summary_sheet(ws2, records, cats, hdr_font, hdr_fill, data_font, bdr, center, left)

    wb.save(out_path)


def export_tabular_xlsx(records: list, meta: dict, out_path: str) -> None:
    """
    Export tabular-source records preserving original columns, appending
    MIU Type / Year / Age analysis columns at the end.
    """
    from collections import Counter
    if not records:
        export_xlsx(records, meta, out_path)
        return

    wb  = Workbook()
    bdr = _border()

    hdr_font  = Font(name="Calibri", bold=True, color=WHITE, size=10)
    hdr_fill  = PatternFill("solid", start_color=NAVY)
    data_font = Font(name="Calibri", size=10)
    analysis_hdr_fill = PatternFill("solid", start_color="2E4057")
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center")

    cats = Counter(r["age_category"] for r in records)

    # ── Sheet 1: MIU Age Report ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "MIU Age Report"

    src_file  = meta.get("source_file", "")
    miu_col   = meta.get("miu_column", "")
    det_meth  = meta.get("detection_method", "")
    det_note  = {
        "header": f"auto-detected from column header \"{miu_col}\"",
        "value":  f"auto-detected by value scan \"{miu_col}\"",
        "user":   f"user-selected \"{miu_col}\"",
        "none":   "MIU column NOT detected — results may be incomplete",
    }.get(det_meth, miu_col)

    # Grab original column headers from first record that has them
    orig_headers = []
    for rec in records:
        if "_headers" in rec:
            orig_headers = rec["_headers"]
            break

    analysis_cols = [
        ("MIU System",    14), ("MIU Type",       26),
        ("Mfr. Year",     12), ("Est. Age",        14),
        ("Age Category",  22),
    ]
    total_cols = len(orig_headers) + len(analysis_cols)
    last_col   = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"MIU System Assessment Tool  --  {src_file}"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        f"Source: {src_file}   |   "
        f"MIU Serial column: {det_note}   |   "
        f"Total rows: {len(records)}"
    )
    ws["A2"].font      = Font(name="Calibri", italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = (
        f"New (<=5 yrs): {cats.get('New (<=5 yrs)', 0)}   |   "
        f"Moderate (6-10 yrs): {cats.get('Moderate (6-10 yrs)', 0)}   |   "
        f"Aging (11-15 yrs): {cats.get('Aging (11-15 yrs)', 0)}   |   "
        f"End of Life (>15 yrs): {cats.get('End of Life (>15 yrs)', 0)}   |   "
        f"Unknown / Invalid: {cats.get('Unknown', 0)}"
    )
    ws["A3"].font      = Font(name="Calibri", bold=True, size=10, color="7B2D00")
    ws["A3"].fill      = PatternFill("solid", start_color="FCE4D6")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 5

    # Header row — original columns + analysis columns
    for ci, col in enumerate(orig_headers, 1):
        c = ws.cell(row=5, column=ci, value=col)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = max(12, min(30, len(str(col)) + 4))
    for offset, (label, width) in enumerate(analysis_cols):
        ci = len(orig_headers) + offset + 1
        c = ws.cell(row=5, column=ci, value=label)
        c.font = hdr_font; c.fill = analysis_hdr_fill; c.alignment = center; c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[5].height = 20

    serial_missing_fill = PatternFill("solid", start_color="FFCCCC")
    serial_invalid_fill = PatternFill("solid", start_color="FFD7A8")
    serial_missing_font = Font(name="Calibri", size=10, bold=True, color="922B21")
    serial_invalid_font = Font(name="Calibri", size=10, bold=True, color="7E3200")

    miu_col_ci = (orig_headers.index(miu_col) + 1) if miu_col in orig_headers else None

    for ri, rec in enumerate(records, 6):
        cat           = rec["age_category"]
        row_fill      = PatternFill("solid", start_color=age_fill_color(cat))
        serial_status = rec.get("serial_status", "ok")
        raw_row       = rec.get("_raw_row", {})

        # Original columns
        for ci, col in enumerate(orig_headers, 1):
            val = raw_row.get(col, "")
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font   = data_font
            c.border = bdr
            c.alignment = center
            if ci == miu_col_ci:
                if serial_status == "missing":
                    c.fill = serial_missing_fill; c.font = serial_missing_font
                elif serial_status == "invalid":
                    c.fill = serial_invalid_fill; c.font = serial_invalid_font
                else:
                    c.fill = row_fill
            else:
                c.fill = row_fill

        # Analysis columns
        analysis_vals = [
            rec["system"],
            rec["miu_type"],
            rec["est_year"] or "--",
            rec["age_label"],
            cat,
        ]
        for offset, val in enumerate(analysis_vals):
            ci = len(orig_headers) + offset + 1
            c  = ws.cell(row=ri, column=ci, value=val)
            c.font = data_font; c.fill = row_fill; c.border = bdr; c.alignment = center
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = f"A6"
    ws.auto_filter.ref = f"A5:{last_col}{len(records) + 5}"

    # ── Sheet 2: Age Summary ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Age Summary")
    _write_summary_sheet(ws2, records, cats, hdr_font, hdr_fill, data_font, bdr, center, left)

    wb.save(out_path)


def _write_summary_sheet(ws2, records, cats, hdr_font, hdr_fill, data_font, bdr, center, left):
    from collections import Counter as _Counter
    ws2["A1"] = "MIU System Assessment Tool  --  Age Summary"
    ws2["A1"].font      = Font(name="Calibri", bold=True, size=13, color=NAVY)
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 24
    ws2.row_dimensions[2].height = 6

    def _tbl_hdr(ws, row, labels, widths):
        for ci, (lbl, w) in enumerate(zip(labels, widths), 1):
            c = ws.cell(row=row, column=ci, value=lbl)
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = bdr
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = 18

    _tbl_hdr(ws2, 3, ["Age Category", "Count", "% of Total"], [24, 10, 14])
    cat_order = ["New (<=5 yrs)", "Moderate (6-10 yrs)",
                 "Aging (11-15 yrs)", "End of Life (>15 yrs)", "Unknown"]
    total = len(records)
    for ri2, cat in enumerate(cat_order, 4):
        count = cats.get(cat, 0)
        pct   = f"{count / total * 100:.1f}%" if total else "--"
        fill  = PatternFill("solid", start_color=age_fill_color(cat))
        for ci, val in enumerate([cat, count, pct], 1):
            c = ws2.cell(row=ri2, column=ci, value=val)
            c.font = data_font; c.fill = fill; c.border = bdr
            c.alignment = left if ci == 1 else center
        ws2.row_dimensions[ri2].height = 16

    ws2.row_dimensions[9].height = 10
    _tbl_hdr(ws2, 10, ["MIU Type", "Count", "% of Total"], [28, 10, 14])
    type_counts = _Counter(r["miu_type"] for r in records)
    alt = [PatternFill("solid", start_color="D6E4F0"),
           PatternFill("solid", start_color="EBF3FB")]
    for ri3, (mtype, count) in enumerate(
            sorted(type_counts.items(), key=lambda x: -x[1]), 11):
        pct  = f"{count / total * 100:.1f}%" if total else "--"
        fill = alt[ri3 % 2]
        for ci, val in enumerate([mtype, count, pct], 1):
            c = ws2.cell(row=ri3, column=ci, value=val)
            c.font = data_font; c.fill = fill; c.border = bdr
            c.alignment = left if ci == 1 else center
        ws2.row_dimensions[ri3].height = 16


# ─── Column Selector Dialog ───────────────────────────────────────────────────

